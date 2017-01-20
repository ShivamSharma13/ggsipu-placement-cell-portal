from student.models import Student
import openpyxl as excel, time
from hashids import Hashids

def get_excel_structure(main_heading, secondary_heading, students_queryset):
	workbook = excel.Workbook()
	worksheet = workbook.active
	worksheet.title = "Placement Session"
	worksheet['A1'].font = excel.styles.Font(name='Times New Roman', size=20, bold=True)
	worksheet.merge_cells("A1:I2");worksheet.merge_cells("A3:I3")
	worksheet['A1'] = main_heading # Job/Internship @ College
	worksheet['A3'] = secondary_heading # Programme - Streams
	worksheet.freeze_panes = 'A5'
	
	# S.No. | Enrollment No. | First Name | Last Name | Gender | Email | Stream | Year
	worksheet['A4'] = 'S.No.'; worksheet['B4'] = 'Enrollment No.'; worksheet['C4'] = 'First Name'; worksheet['D4'] = 'Last Name';
	worksheet['E4'] = 'Sex';worksheet['F4'] = 'Email'; worksheet['G4'] = 'Stream'; worksheet['H4'] = 'Year';
	
	bold = excel.styles.Font(bold=True)
	for i in range(1,9):
		worksheet.cell(row=4, column=i).font = bold
	
	GENDER = dict(Student.GENDER_CHOICES)
	for i, student in enumerate(students_queryset, 1):
		row = worksheet.max_row+1
		worksheet.cell(row=row, column=1).value = i
		worksheet.cell(row=row, column=2).value = student.profile.username
		worksheet.cell(row=row, column=3).value = student.firstname.title()
		worksheet.cell(row=row, column=4).value = student.lastname.title()
		worksheet.cell(row=row, column=5).value = GENDER[student.gender].__str__()
		worksheet.cell(row=row, column=6).value = student.profile.email
		worksheet.cell(row=row, column=7).value = student.stream.name.title()
		worksheet.cell(row=row, column=8).value = student.current_year
	
	to_letter = excel.cell.get_column_letter
	for col in [1,5,8]:
		worksheet.column_dimensions[to_letter(col)].width = 5
	for col in [2,3,4]:
		worksheet.column_dimensions[to_letter(col)].width = 12
	for col in [6,7]:
		worksheet.column_dimensions[to_letter(col)].width = 20
	return workbook
